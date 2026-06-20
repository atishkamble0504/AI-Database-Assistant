import sqlite3
from openai import OpenAI
import os
from dotenv import load_dotenv
import re
import json
import csv
import io
from openpyxl import Workbook

# -----------------------------
# Load API key (OpenRouter)
# -----------------------------
load_dotenv()

client = OpenAI(
    api_key=os.getenv("OPENAI_API_KEY"),
    base_url="https://openrouter.ai/api/v1",
    default_headers={
        "HTTP-Referer": "http://localhost",
        "X-Title": "AI DB Demo"
    }
)

UPLOAD_FOLDER = "uploads"

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)


def save_uploaded_file(file):
    if file and file.filename.endswith(".db"):
        file_path = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(file_path)
        return file_path
    return None


# -----------------------------
# SCHEMA
# -----------------------------
def get_schema(db_path):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
    tables = cursor.fetchall()

    schema = ""

    for table in tables:
        table_name = table[0]

        if table_name.startswith("sqlite_"):
            continue

        cursor.execute(f"PRAGMA table_info({table_name})")
        columns = cursor.fetchall()
        col_names = [col[1] for col in columns]

        schema += f"{table_name}({', '.join(col_names)})\n"

    conn.close()
    return schema


# -----------------------------
# AI → SQL (Enhanced for Auto-Correction)
# -----------------------------
def nl_to_sql(user_input, schema, error_context=None, conversation_history=None):
    prompt = f"""
You are a smart AI that converts English into SQL.

Database schema:
{schema}
"""

    # 🔥 ADD CONTEXT HERE
    if conversation_history:
        prompt += "\nPrevious conversation:\n"
        for item in conversation_history[-3:]:  # last 3 queries only
            prompt += f"- {item['user']} → {item['sql']}\n"

    if error_context:
        prompt += f"\nSTRICT: The previous query failed with error: {error_context}. Fix the SQL syntax or table/column names."

    prompt += f"""
STRICT RULES:
- Use ONLY tables and columns from schema when querying
- You ARE allowed to CREATE and DROP tables
- Use LOWER(column) = LOWER(value) for text
- SQLite syntax only
- Allowed: SELECT, INSERT, UPDATE, DELETE, CREATE, DROP
- No explanation

Convert:
{user_input}
"""

    response = client.chat.completions.create(
        model="openrouter/auto",
        messages=[{"role": "user", "content": prompt}]
    )

    sql = response.choices[0].message.content.strip()
    sql = sql.replace("```sql", "").replace("```", "").strip()

    if sql.strip().upper().startswith("SELECT"):
        sql = sql.upper().split("SELECT", 1)[1]
        sql = "SELECT " + sql

    return sql

# -----------------------------
# 🧠 SQL Explanation
# -----------------------------
def explain_sql(sql):
    prompt = f"""
Explain this SQL query in simple English.

SQL:
{sql}

Keep it short and clear.
"""

    try:
        response = client.chat.completions.create(
            model="openrouter/auto",
            messages=[{"role": "user", "content": prompt}]
        )

        explanation = response.choices[0].message.content.strip()
        return explanation

    except:
        return "Could not generate explanation"


# -----------------------------
# Case-insensitive SQL
# -----------------------------
def make_case_insensitive(sql):
    pattern = r"(\w+)\s*=\s*'([^']+)'"

    def repl(match):
        col = match.group(1)
        val = match.group(2)
        return f"LOWER({col}) = LOWER('{val}')"

    return re.sub(pattern, repl, sql)


# -----------------------------
# 🔥 DANGER DETECTION ENGINE
# -----------------------------
def detect_dangerous_query(sql):
    sql_lower = sql.lower()

    if "drop table" in sql_lower:
        match = re.search(r"drop table (\w+)", sql_lower)
        table = match.group(1) if match else "this table"

        return {
            "type": "warning",
            "message": f"⚠️ This will permanently DELETE table '{table}' and ALL its data. This cannot be undone."
        }

    if sql_lower.startswith("delete") and "where" not in sql_lower:
        match = re.search(r"from (\w+)", sql_lower)
        table = match.group(1) if match else "this table"

        return {
            "type": "warning",
            "message": f"⚠️ This will DELETE ALL records from '{table}' table. All data will be lost."
        }

    if sql_lower.startswith("update") and "where" not in sql_lower:
        match = re.search(r"update (\w+)", sql_lower)
        table = match.group(1) if match else "this table"

        return {
            "type": "warning",
            "message": f"⚠️ This will UPDATE ALL rows in '{table}' table. Previous values will be lost."
        }

    return None


# -----------------------------
# Suggestion Engine
# -----------------------------
def get_suggestions(db_path, sql):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    suggestions = []

    try:
        match = re.search(r"FROM\s+(\w+)", sql, re.IGNORECASE)
        if not match:
            return []

        table = match.group(1)

        cursor.execute(f"PRAGMA table_info({table})")
        columns = cursor.fetchall()

        for col in columns:
            col_name = col[1]

            cursor.execute(f"SELECT DISTINCT {col_name} FROM {table} LIMIT 20")
            values = cursor.fetchall()

            for v in values:
                if isinstance(v[0], str):
                    suggestions.append(v[0])

        conn.close()
        return list(set(suggestions))[:5]

    except:
        conn.close()
        return []


# ---------------------------------------------------------
# 🚀 SQL Execution (Updated with Auto-Validation Layer)
# ---------------------------------------------------------
def run_query(db_path, sql, user_input=None, confirm=False, retry=True):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    sql_lower = sql.lower().strip()

    danger = detect_dangerous_query(sql)

    if danger and not confirm:
        conn.close()
        return {
            "type": "warning",
            "message": danger["message"],
            "sql": sql
        }

    try:
        sql = make_case_insensitive(sql)
        cursor.execute(sql)

        if sql_lower.startswith("select"):
            columns = [desc[0] for desc in cursor.description]
            results = cursor.fetchall()

            if len(results) == 0:
                suggestions = get_suggestions(db_path, sql)
                conn.close()
                if suggestions:
                    return {
                        "type": "suggestion",
                        "message": "⚠️ No exact match. Try these:",
                        "suggestions": suggestions
                    }
                return {"type": "empty", "message": "❌ No data available"}

            conn.close()
            return {
    "type": "exact",
    "data": results,
    "columns": columns,
    "explanation": explain_sql(sql)   # 🔥 NEW
}

        if sql_lower.startswith("create"):
            conn.commit()
            conn.close()
            return {"type": "success", "message": "🆕 Table created successfully"}

        conn.commit()
        conn.close()
        return {
    "type": "success",
    "message": "✅ Query executed successfully",
    "explanation": explain_sql(sql)
}

    except Exception as e:
        conn.close()
        # LAYER 3: If SQL fails, re-try once with the error message
        if retry and user_input:
            current_schema = get_schema(db_path)
            new_sql = nl_to_sql(user_input, current_schema, error_context=str(e))
            return run_query(db_path, new_sql, user_input, confirm, retry=False)
            
        return {"type": "error", "message": f"❌ SQL Error: {e}"}


# =========================================================
# 🚀 EXPORT SYSTEM
# =========================================================

def result_to_csv(data, columns):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(columns)
    writer.writerows(data)
    return output.getvalue()


def result_to_json(data, columns):
    result = [dict(zip(columns, row)) for row in data]
    return json.dumps(result, indent=2)


def result_to_excel(data, columns):
    wb = Workbook()
    ws = wb.active
    ws.append(columns)

    for row in data:
        ws.append(row)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def get_all_tables(db_path):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    cursor.execute("""
        SELECT name 
        FROM sqlite_master 
        WHERE type='table' 
        AND name NOT LIKE 'sqlite_%';
    """)

    tables = [row[0] for row in cursor.fetchall()]

    conn.close()
    return tables


def export_db_json(db_path):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    data = {}
    tables = get_all_tables(db_path)

    for table in tables:
        cursor.execute(f"SELECT * FROM {table}")
        columns = [d[0] for d in cursor.description]
        rows = cursor.fetchall()

        data[table] = [dict(zip(columns, r)) for r in rows]

    conn.close()
    return json.dumps(data, indent=2)


def export_db_excel(db_path):
    wb = Workbook()
    wb.remove(wb.active)

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    tables = get_all_tables(db_path)

    for table in tables:
        ws = wb.create_sheet(title=table)

        cursor.execute(f"SELECT * FROM {table}")
        columns = [d[0] for d in cursor.description]
        rows = cursor.fetchall()

        ws.append(columns)
        for r in rows:
            ws.append(r)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    conn.close()
    return stream


def export_single_table_csv(db_path, table):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    cursor.execute(f"SELECT * FROM {table}")
    columns = [d[0] for d in cursor.description]
    rows = cursor.fetchall()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(columns)
    writer.writerows(rows)

    conn.close()
    return output.getvalue()


# -----------------------------
# DB Resolver
# -----------------------------
def resolve_db_path(text_path, uploaded_file):
    if uploaded_file:
        saved_path = save_uploaded_file(uploaded_file)
        if saved_path:
            return saved_path

    if text_path:
        return text_path

    return None


# -----------------------------
# CLI TEST
# -----------------------------
def main():
    print("🚀 Smart AI Database (CLI Mode)\n")

    db_path = input("Enter DB path: ")
    schema = get_schema(db_path)

    while True:
        user_input = input("\nAsk: ")

        if user_input.lower() == "exit":
            break

        sql = nl_to_sql(user_input, schema)
        print("\n🧠 SQL:", sql)

        result = run_query(db_path, sql)
        print("\n📊 Result:", result)


if __name__ == "__main__":
    main()